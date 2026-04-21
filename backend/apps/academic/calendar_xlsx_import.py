"""
Importação do calendário escolar a partir da planilha padrão (layout 2026).

Layout esperado: uma aba com blocos de meses nas colunas B, K, T, AC, AL, AU;
linhas 14–21 (1.º semestre) e 31–37 (2.º semestre) com textos do tipo
"Dia - descrição" ou "Dia a Dia - descrição".
"""
from __future__ import annotations

import calendar
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Iterator
from zoneinfo import ZoneInfo

MONTH_NAMES_PT = frozenset(
    x.lower()
    for x in (
        'Janeiro',
        'Fevereiro',
        'Março',
        'Marco',
        'Abril',
        'Maio',
        'Junho',
        'Julho',
        'Agosto',
        'Setembro',
        'Outubro',
        'Novembro',
        'Dezembro',
    )
)

# Colunas dos blocos de mês (openpyxl / Excel)
MONTH_COLS = ('B', 'K', 'T', 'AC', 'AL', 'AU')

TZ_SP = ZoneInfo('America/Sao_Paulo')

IMPORT_MARKER = '[Fonte: importação calendário xlsx]'
IMPORT_DESCRIPTION_SUFFIX = f'\n\n{IMPORT_MARKER}'


@dataclass
class ParsedCalendarEvent:
    start_date: date
    end_date: date
    description: str
    title: str
    event_type: str
    target_audience: str


def _norm(s: str) -> str:
    s = unicodedata.normalize('NFKC', s or '')
    s = s.replace('–', '-').replace('—', '-').strip()
    s = re.sub(r'\s+', ' ', s)
    return s


def _should_skip_cell(text: str) -> bool:
    t = text.lower()
    if not t:
        return True
    if t in MONTH_NAMES_PT:
        return True
    if re.fullmatch(r'[dstq]', t):
        return True
    if 'dia letivo' in t and 'dias letivos' not in t:
        return True
    if re.search(r'\d+\s+dias?\s+letivos', t):
        return True
    if 'obs.:' in t or t.startswith('legenda'):
        return True
    return False


def classify_event(description: str) -> tuple[str, str]:
    """Retorna (event_type, target_audience) para SchoolEvent."""
    t = description.lower()
    if 'capacitação' in t or 'capacitacao' in t:
        return 'MEETING', 'TEACHERS'
    if 'reunião' in t or 'reuniao' in t:
        return 'MEETING', 'ALL'
    if (
        'feriado' in t
        or 'recesso' in t
        or 'suspensão' in t
        or 'suspensao' in t
        or 'emenda' in t
        or 'férias escolares' in t
        or 'ferias escolares' in t
        or 'aulas suspensas' in t
    ):
        return 'HOLIDAY', 'ALL'
    if (
        'bimestre' in t
        or 'início do ano letivo' in t
        or 'inicio do ano letivo' in t
        or 'início das aulas' in t
        or 'inicio das aulas' in t
        or 'término do' in t
        or 'termino do' in t
        or 'início do' in t
        or 'inicio do' in t
    ):
        return 'SCHOOL_DAY', 'ALL'
    if 'colônia' in t or 'colonia' in t or 'curso de férias' in t or 'curso de ferias' in t:
        return 'EVENT', 'ALL'
    return 'EVENT', 'ALL'


def _make_title(rest: str, max_len: int = 200) -> str:
    rest = rest.strip()
    if len(rest) <= max_len:
        return rest
    return rest[: max_len - 1] + '…'


def _safe_date(year: int, month: int, day: int) -> date | None:
    last = calendar.monthrange(year, month)[1]
    if day < 1 or day > last:
        return None
    return date(year, month, day)


def parse_calendar_line(raw: str, year: int, month: int) -> list[ParsedCalendarEvent]:
    """
    Interpreta uma célula de evento do calendário.
    Devolve zero ou mais ParsedCalendarEvent (várias se o texto não couber num único intervalo válido).
    """
    text = _norm(str(raw))
    if _should_skip_cell(text):
        return []

    # Intervalo: "02 a 14 - Recesso"
    m = re.match(r'^(\d{1,2})\s+a\s+(\d{1,2})\s*-\s*(.+)$', text, re.IGNORECASE)
    if m:
        d1, d2 = int(m.group(1)), int(m.group(2))
        rest = m.group(3).strip()
        sd = _safe_date(year, month, d1)
        ed = _safe_date(year, month, d2)
        if not sd or not ed or sd > ed:
            return []
        et, aud = classify_event(text)
        return [
            ParsedCalendarEvent(
                start_date=sd,
                end_date=ed,
                description=text,
                title=_make_title(rest),
                event_type=et,
                target_audience=aud,
            )
        ]

    # "16, 17 e 18 - Recesso"
    m = re.match(r'^(\d{1,2}),\s*(\d{1,2})\s+e\s+(\d{1,2})\s*-\s*(.+)$', text, re.IGNORECASE)
    if m:
        days = sorted(int(m.group(i)) for i in range(1, 4))
        rest = m.group(4).strip()
        sd = _safe_date(year, month, days[0])
        ed = _safe_date(year, month, days[2])
        if not sd or not ed:
            return []
        et, aud = classify_event(text)
        return [
            ParsedCalendarEvent(
                start_date=sd,
                end_date=ed,
                description=text,
                title=_make_title(rest),
                event_type=et,
                target_audience=aud,
            )
        ]

    # Dia único: "31 - Reunião de Pais" ou "6 - Atividade Interna"
    m = re.match(r'^(\d{1,2})\s*-\s*(.+)$', text, re.IGNORECASE)
    if m:
        d = int(m.group(1))
        rest = m.group(2).strip()
        sd = _safe_date(year, month, d)
        if not sd:
            return []
        et, aud = classify_event(text)
        return [
            ParsedCalendarEvent(
                start_date=sd,
                end_date=sd,
                description=text,
                title=_make_title(rest),
                event_type=et,
                target_audience=aud,
            )
        ]

    return []


def _cell_str(value) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def iter_parsed_events(
    xlsx_path: str | Path,
    year: int = 2026,
    sheet_name: str | None = None,
) -> Iterator[ParsedCalendarEvent]:
    """
    Lê o workbook e gera eventos interpretados.
    """
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError(
            'Instale openpyxl para usar a importação do calendário (pip install openpyxl).'
        ) from e

    path = Path(xlsx_path)
    # Sem read_only: o modo read-only não permite ws['B14'] (acesso aleatório).
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.worksheets[0]
        # 1.º semestre: meses 1–6, linhas 14–21
        for row in range(14, 22):
            for idx, col in enumerate(MONTH_COLS):
                month = idx + 1
                val = _cell_str(ws[f'{col}{row}'].value)
                for ev in parse_calendar_line(val, year, month):
                    yield ev
        # 2.º semestre: meses 7–12, linhas 31–37
        for row in range(31, 38):
            for idx, col in enumerate(MONTH_COLS):
                month = idx + 7
                val = _cell_str(ws[f'{col}{row}'].value)
                for ev in parse_calendar_line(val, year, month):
                    yield ev
    finally:
        wb.close()


def _to_aware_start(d: date) -> datetime:
    return datetime.combine(d, time(8, 0), tzinfo=TZ_SP)


def _to_aware_end(d: date) -> datetime:
    return datetime.combine(d, time(17, 0), tzinfo=TZ_SP)


def import_calendar_events(
    xlsx_path: str | Path,
    *,
    year: int = 2026,
    dry_run: bool = False,
    skip_existing: bool = False,
    created_by=None,
    stdout=None,
) -> dict:
    """
    Persiste SchoolEvent a partir do xlsx. Retorna contadores.
    """
    from django.db import transaction

    from apps.academic.models import SchoolEvent

    def log(msg: str):
        if stdout:
            stdout.write(msg)

    created = 0
    dry_run_lines = 0
    skipped_dup = 0

    events = list(iter_parsed_events(xlsx_path, year=year))

    def persist():
        nonlocal created, dry_run_lines, skipped_dup
        for ev in events:
            desc = ev.description + IMPORT_DESCRIPTION_SUFFIX

            if skip_existing and SchoolEvent.objects.filter(
                title=ev.title,
                start_time__date=ev.start_date,
            ).exists():
                skipped_dup += 1
                continue

            if dry_run:
                log(
                    f'[dry-run] {ev.start_date} … {ev.end_date} | {ev.event_type}/{ev.target_audience} | {ev.title[:80]}\n'
                )
                dry_run_lines += 1
                continue

            SchoolEvent.objects.create(
                title=ev.title,
                description=desc,
                start_time=_to_aware_start(ev.start_date),
                end_time=_to_aware_end(ev.end_date),
                event_type=ev.event_type,
                target_audience=ev.target_audience,
                classroom=None,
                subject=None,
                created_by=created_by,
            )
            created += 1

    if dry_run:
        persist()
    else:
        with transaction.atomic():
            persist()

    return {
        'parsed': len(events),
        'created': created,
        'dry_run_lines': dry_run_lines,
        'skipped_duplicate': skipped_dup,
    }
